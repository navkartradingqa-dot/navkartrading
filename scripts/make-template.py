"""
Builds public/navkar-product-template.xlsx — the blank sheet shop staff fill in
and upload at Admin → Import. Categories and brands come from the live
catalogue so the dropdowns can only offer things that already exist.

Run:  python3 scripts/make-template.py
"""

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

CATEGORIES = [
    "Laptops",
    "Mobile Phones",
    "Tablets",
    "Mobile Accessories",
    "Audio & Headphones",
    "Smart Watches",
    "Computer Accessories",
    "Monitors & Displays",
    "Storage & Memory",
    "Networking",
    "Gaming",
    "Cameras & Drones",
    "Printers & Scanners",
    "Power & Charging",
    "Smart Home",
]

BRANDS = [
    "ASUS", "Acer", "Anker", "Apple", "Baseus", "Belkin", "Brother", "Canon",
    "Corsair", "DJI", "Dell", "Epson", "GoPro", "Google", "HP", "Honor",
    "Huawei", "JBL", "Kingston", "LG", "Lenovo", "Lexar", "Logitech", "MSI",
    "Netgear", "Nothing", "OnePlus", "Philips", "Razer", "Samsung", "SanDisk",
    "Seagate", "Sony", "TP-Link", "Tecno", "UGREEN", "WD", "Xiaomi",
]

# (heading, width, help text shown on the instructions sheet)
COLUMNS = [
    ("SKU", 16, "Required. Your own code for the item — also what the barcode scanner reads. Same SKU = update that product, new SKU = new product."),
    ("Name (EN)", 34, "Required for a new product. What customers see."),
    ("Name (AR)", 26, "Optional. Leave blank and the English name is used on the Arabic site."),
    ("Category", 20, "Required for a new product. Pick from the dropdown."),
    ("Brand", 14, "Optional. A brand that does not exist yet is created for you."),
    ("Price", 11, "Required for a new product. The normal selling price in QAR, before any discount."),
    ("Discount %", 11, "Optional. 10 means 10% off. The site shows the old price struck through."),
    ("Cost", 11, "Optional. What you paid. Used for margin reports, never shown to customers."),
    ("Stock", 9, "How many you have. Changing this is logged as a stock movement."),
    ("Reorder at", 11, "Optional. Low-stock warning level. Default 5."),
    ("Warranty (months)", 17, "Optional. Default 12."),
    ("Barcode", 16, "Optional. The manufacturer's EAN/UPC, if the box has one."),
    ("Description (EN)", 42, "Optional. A short paragraph for the product page."),
    ("Description (AR)", 30, "Optional."),
    ("Specs", 34, "Optional. One per line, 'Key: value' — e.g. Storage: 512GB"),
    ("Active", 9, "yes or no. Should it show on the website? Default yes."),
    ("Featured", 10, "yes or no. Show on the home page? Default no."),
]

EXAMPLES = [
    ["NT-LAP-0101", "HP Pavilion 15 — Core i5 / 16GB / 512GB", "", "Laptops", "HP",
     3299, 10, 2750, 6, 2, 24, "195161234567", "Everyday laptop for study and office work.",
     "", "Processor: Core i5-1335U\nMemory: 16GB\nStorage: 512GB SSD", "yes", "yes"],
    ["NT-MOB-0102", "Samsung Galaxy A55 5G 256GB", "سامسونج جالاكسي A55", "Mobile Phones",
     "Samsung", 1799, "", 1450, 10, 3, 12, "", "Dual SIM, local warranty.", "",
     "Display: 6.6\" AMOLED\nStorage: 256GB", "yes", "no"],
    ["NT-AUD-0103", "JBL Tune 770NC Wireless Headphones", "", "Audio & Headphones", "JBL",
     549, 15, 400, 14, 5, 12, "", "Noise cancelling, 70-hour battery.", "",
     "Battery: 70 hours\nConnection: Bluetooth 5.3", "yes", "no"],
]

HEADER_FILL = PatternFill("solid", fgColor="8F1F43")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
REQUIRED_FILL = PatternFill("solid", fgColor="FBE5E8")
THIN = Side(style="thin", color="DADCE0")


def build(path: str) -> None:
    wb = openpyxl.Workbook()

    # ---------------------------------------------------------- Products
    ws = wb.active
    ws.title = "Products"

    for i, (heading, width, _) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=i, value=heading)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(i)].width = width
    ws.row_dimensions[1].height = 30

    for row in EXAMPLES:
        ws.append(row)

    for r in range(2, 2 + len(EXAMPLES)):
        for c in range(1, len(COLUMNS) + 1):
            cell = ws.cell(row=r, column=c)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = Border(bottom=THIN)
        # SKU, Name, Category and Price are the ones you cannot leave out.
        for c in (1, 2, 4, 6):
            ws.cell(row=r, column=c).fill = REQUIRED_FILL

    ws.freeze_panes = "A2"

    # ------------------------------------------------------------- Lists
    lists = wb.create_sheet("Lists")
    lists["A1"] = "Categories"
    lists["B1"] = "Brands"
    lists["C1"] = "Yes / no"
    for i, name in enumerate(CATEGORIES, start=2):
        lists.cell(row=i, column=1, value=name)
    for i, name in enumerate(BRANDS, start=2):
        lists.cell(row=i, column=2, value=name)
    lists["C2"] = "yes"
    lists["C3"] = "no"
    lists.column_dimensions["A"].width = 24
    lists.column_dimensions["B"].width = 18
    lists.sheet_state = "hidden"

    last = 500  # dropdowns apply this far down so pasted rows are covered

    category_dv = DataValidation(
        type="list",
        formula1=f"=Lists!$A$2:$A${len(CATEGORIES) + 1}",
        allow_blank=True,
        showDropDown=False,
    )
    category_dv.error = "Pick a category from the list, or add the new category in Admin → Settings first."
    category_dv.errorTitle = "Unknown category"
    ws.add_data_validation(category_dv)
    category_dv.add(f"D2:D{last}")

    brand_dv = DataValidation(
        type="list",
        formula1=f"=Lists!$B$2:$B${len(BRANDS) + 1}",
        allow_blank=True,
        showDropDown=False,
    )
    brand_dv.errorStyle = "warning"
    brand_dv.error = "That brand does not exist yet — it will be created when you import."
    brand_dv.errorTitle = "New brand"
    ws.add_data_validation(brand_dv)
    brand_dv.add(f"E2:E{last}")

    yesno_dv = DataValidation(
        type="list", formula1="=Lists!$C$2:$C$3", allow_blank=True, showDropDown=False
    )
    ws.add_data_validation(yesno_dv)
    yesno_dv.add(f"P2:P{last}")
    yesno_dv.add(f"Q2:Q{last}")

    discount_dv = DataValidation(
        type="decimal", operator="between", formula1=0, formula2=99, allow_blank=True
    )
    discount_dv.error = "Enter the discount as a number of percent, e.g. 10 for 10% off."
    discount_dv.errorTitle = "Discount %"
    ws.add_data_validation(discount_dv)
    discount_dv.add(f"G2:G{last}")

    # ------------------------------------------------------ How to use
    guide = wb.create_sheet("How to use", 0)
    guide.column_dimensions["A"].width = 22
    guide.column_dimensions["B"].width = 96

    def line(label: str, text: str = "", bold: bool = False, gap: bool = False):
        r = guide.max_row + (2 if gap else 1)
        if guide.max_row == 1 and guide["A1"].value is None:
            r = 1
        guide.cell(row=r, column=1, value=label).font = Font(bold=True, size=12 if bold else 11)
        c = guide.cell(row=r, column=2, value=text)
        c.alignment = Alignment(wrap_text=True, vertical="top")

    line("Navkar Trading", "Product import template", bold=True)
    line("", "")
    line("1.", "Fill in the Products sheet — one row per product. The three example rows are there to show the format; delete them before you import.")
    line("2.", "Go to the website's admin panel → Import, and upload this file.")
    line("3.", "You get a summary of exactly what will be added and changed. Nothing is saved until you press Apply.")
    line("", "")
    line("The golden rule", "SKU is what links a row to a product. A SKU that already exists updates that product. A SKU that does not exist creates a new one. Never reuse a SKU for a different item.", gap=False)
    line("", "")
    line("Only fill what you need", "Leave a whole column out and that field is left alone. To change only prices, delete every column except SKU and Price.")
    line("", "")
    line("Photos", "Do not put photos in this file. Name each photo after its SKU — NT-LAP-0101.jpg — and drop them on the Photos tab of the same Import page. For a second photo of the same product use NT-LAP-0101-2.jpg.")
    line("", "")
    line("Updating what you already have", "On the Import page, click 'Export current products' first. Edit that file in Excel and import it back — the SKUs and category names will already be correct.")
    line("", "")
    line("Columns", "", bold=True)
    for heading, _, help_text in COLUMNS:
        line(heading, help_text)

    guide.sheet_view.showGridLines = False

    wb.active = 0
    wb.save(path)


if __name__ == "__main__":
    build("public/navkar-product-template.xlsx")
    print("wrote public/navkar-product-template.xlsx")
